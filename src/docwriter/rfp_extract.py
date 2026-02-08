from __future__ import annotations

from dataclasses import dataclass
import io
import json
import string
from typing import Dict, List, Tuple

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None  # type: ignore

try:
    from docx import Document  # type: ignore
except Exception:  # pragma: no cover
    Document = None  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None  # type: ignore


def _normalize_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines()]
    cleaned = [line for line in lines if line]
    return "\n".join(cleaned)


def extract_pdf_text(data_bytes: bytes) -> str:
    if PdfReader is None:
        raise RuntimeError("pypdf is required to extract PDF text")
    reader = PdfReader(io.BytesIO(data_bytes))
    parts: List[str] = []
    for page in reader.pages:
        page_text = page.extract_text() or ""
        if page_text.strip():
            parts.append(page_text)
    return _normalize_text("\n".join(parts))


def extract_docx_text(data_bytes: bytes) -> str:
    if Document is None:
        raise RuntimeError("python-docx is required to extract DOCX text")
    doc = Document(io.BytesIO(data_bytes))
    parts = [p.text for p in doc.paragraphs if p.text and p.text.strip()]
    return _normalize_text("\n".join(parts))


@dataclass
class XlsxExtraction:
    text: str
    json_payload: Dict[str, List[Dict[str, str]]]


def _column_letters(count: int) -> List[str]:
    letters = []
    alphabet = string.ascii_uppercase
    for idx in range(count):
        first = idx // 26
        second = idx % 26
        prefix = alphabet[first - 1] if first else ""
        letters.append(f"{prefix}{alphabet[second]}")
    return letters


def extract_xlsx_text_and_json(data_bytes: bytes) -> XlsxExtraction:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to extract XLSX text")
    wb = load_workbook(io.BytesIO(data_bytes), data_only=True)
    text_parts: List[str] = []
    json_payload: Dict[str, List[Dict[str, str]]] = {}
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        text_parts.append(f"=== SHEET: {sheet.title} ===")
        header_row = rows[0]
        header_values = [str(cell).strip() if cell is not None else "" for cell in header_row]
        has_header = any(header_values)
        if not has_header:
            header_values = _column_letters(len(header_row))
        json_rows: List[Dict[str, str]] = []
        for row in rows[1 if has_header else 0 :]:
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if not any(values):
                continue
            row_text = " | ".join(values)
            text_parts.append(row_text)
            row_dict = {
                header_values[idx] if idx < len(header_values) else f"COL{idx+1}": value
                for idx, value in enumerate(values)
            }
            json_rows.append(row_dict)
        if json_rows:
            json_payload[sheet.title] = json_rows
    return XlsxExtraction(text=_normalize_text("\n".join(text_parts)), json_payload=json_payload)


def xlsx_json_to_text(json_payload: Dict[str, List[Dict[str, str]]]) -> str:
    return json.dumps(json_payload, indent=2)
